#!/usr/bin/env python3
"""
parse_inventory_v3.py — 새 구조 (2026-05-02~)

지침:
1. 차종(모델명) 중심 — Virtual/Actual 합산이 메인
2. Virtual 비율은 "대기수요 인기도" 시그널로 함께 노출
3. G클래스 완전 제외 (메인 트렌드에서 빠짐)
4. 위탁재고 + 전시차재고 = 일반 재고와 완전 합산
5. 모터원 출고 데이터 무시
6. 타파트너사 해약 감지는 백엔드만 (UI 노출 X)
7. 일/주/월 단위 전국 계약 트렌드가 핵심

Virtual VIN 식별: 차대번호가 'DK'로 시작 OR 차량용도='Virtual VIN' OR 프로세스유형='Dummy'
Actual VIN: 차대번호가 'W1'로 시작 (정식 WMI)
G클래스 제외: 모델명에 ' G '로 시작 또는 'AMG G '로 시작 (단, 'GT'는 제외)
"""

import openpyxl
import re
import json
from pathlib import Path
from datetime import datetime


# ============== 카테고리 분류 (G클래스 분리, 메인은 G 제외) ==============
def categorize(model_name):
    """모델명 → 카테고리 분류. G클래스는 별도 표시."""
    m = model_name
    if 'Maybach' in m: return 'Maybach'
    if m.startswith('EQ'): return '전동화 EQ'  # EQA, EQB, EQE, EQS만
    if 'GLE' in m: return 'GLE'
    if 'GLS' in m: return 'GLS'
    if 'GLC' in m: return 'GLC'
    if 'GLA' in m or 'GLB' in m: return 'GLA·GLB'
    if 'GT ' in m and 'AMG' in m: return 'AMG GT·SL'
    if m.startswith('AMG SL'): return 'AMG GT·SL'
    # G클래스: G로 시작하지만 GLA/GLB/GLC/GLE/GLS/GT는 위에서 처리됨
    if m.startswith('G ') or (m.startswith('AMG G') and 'GT' not in m and 'GLC' not in m and 'GLE' not in m):
        return 'G클래스'
    if m.startswith('S ') or 'AMG S ' in m: return 'S클래스'
    if 'CLE' in m: return 'CLE'
    if m.startswith('E ') or 'AMG E ' in m: return 'E클래스'
    if m.startswith('C ') and 'CL' not in m: return 'C클래스'
    if 'CLA' in m: return 'CLA'
    if m.startswith('A ') or 'AMG A ' in m: return 'A클래스'
    return '기타'


def is_g_class(model_name):
    """G클래스 여부."""
    return categorize(model_name) == 'G클래스'


def clean_model(name):
    """모델명 정규화 — 이중 공백 제거, 'Mercedes-AMG' → 'AMG' 통일."""
    if not name: return ''
    s = re.sub(r'\s+', ' ', str(name)).strip()
    # Mercedes-AMG는 AMG로 통일 (G 63 등 카테고리 매칭용)
    s = re.sub(r'^Mercedes-AMG\s+', 'AMG ', s)
    s = re.sub(r'\bMercedes-AMG\b', 'AMG', s)
    return s


def is_virtual_vin(vin, vehicle_purpose, process_type):
    """Virtual VIN 판별. 'DK'로 시작 OR purpose='Virtual VIN' OR process='Dummy'"""
    if vin and str(vin).startswith('DK'): return True
    if vehicle_purpose and 'Virtual' in str(vehicle_purpose): return True
    if process_type and str(process_type).strip() == 'Dummy': return True
    return False


# ============== allocation 시트 파싱 ==============
def find_col(header, names):
    """헤더에서 컬럼명에 매칭되는 인덱스 찾기 (여러 후보)."""
    for cand in names:
        for i, h in enumerate(header):
            if h and cand in str(h):
                return i
    return -1


def parse_excel(filepath):
    """엑셀 파일 → 표준 레코드 리스트.

    리턴: dict
      {
        "date": "2026-04-30",
        "rows": [{vin, com, model, model_code, ext_color, int_color,
                  pdd, prod_date, branch, salesman, customer,
                  inv_class, car_status, sale_status, vehicle_purpose,
                  process_type, is_virtual, source_sheet}, ...]
      }
    """
    fp = Path(filepath)
    # 파일명에서 날짜 추출
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fp.name)
    date_str = date_match.group(1) if date_match else None

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    rows = []

    # ----- allocation 시트 -----
    if 'allocation' in wb.sheetnames:
        ws = wb['allocation']
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = list(row)
                continue
            if not any(row): continue

            i_co = find_col(header, ['CO.'])
            i_class = find_col(header, ['재고구분'])
            i_branch = find_col(header, ['배정 전시장', '배정전시장'])
            i_team = find_col(header, ['팀'])
            i_sales = find_col(header, ['Salesmen', 'Sales'])
            i_cust = find_col(header, ['COSTOMER', 'CUSTOMER', '고객'])
            i_pdd = find_col(header, ['차량 출고 가능일(PDD)', 'PDD', '출고 가능일'])
            i_com = find_col(header, ['커미션 번호', '커미션'])
            i_mcode = find_col(header, ['모델 코드', '모델코드'])
            i_mname = find_col(header, ['모델명', 'DFE모델명'])
            i_vin = find_col(header, ['차대 번호', '차대번호'])
            i_ext = find_col(header, ['외장 색상'])
            i_int = find_col(header, ['내장 색상'])
            i_prod = find_col(header, ['생산 일자'])
            i_proc = find_col(header, ['프로세스 유형'])
            i_purp = find_col(header, ['차량 용도'])
            i_unsold = find_col(header, ['미판매 재고'])
            i_status = find_col(header, ['재고 상태'])
            i_carstat = find_col(header, ['차량 상태'])
            i_salestat = find_col(header, ['판매 상태'])
            break

        # 다시 처음부터 데이터 읽기
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            if not any(row): continue

            def g(idx):
                return row[idx] if 0 <= idx < len(row) else None

            mname = clean_model(g(i_mname))
            if not mname: continue
            vin = g(i_vin)
            com = g(i_com)
            purpose = g(i_purp)
            proctype = g(i_proc)

            rows.append({
                'date': date_str,
                'vin': str(vin).strip() if vin else None,
                'com': str(com).strip() if com else None,
                'model': mname,
                'model_code': str(g(i_mcode)).strip() if g(i_mcode) else None,
                'ext_color': clean_model(g(i_ext)),
                'int_color': clean_model(g(i_int)),
                'pdd': str(g(i_pdd)) if g(i_pdd) else None,
                'prod_date': str(g(i_prod)) if g(i_prod) else None,
                'branch': clean_model(g(i_branch)),
                'team': clean_model(g(i_team)),
                'salesman': clean_model(g(i_sales)),
                'customer': clean_model(g(i_cust)),
                'inv_class': clean_model(g(i_class)),
                'car_status': clean_model(g(i_carstat)),
                'sale_status': clean_model(g(i_salestat)),
                'stock_status': clean_model(g(i_status)),
                'vehicle_purpose': clean_model(purpose),
                'process_type': clean_model(proctype),
                'unsold': clean_model(g(i_unsold)),
                'is_virtual': is_virtual_vin(vin, purpose, proctype),
                'source': 'allocation',
            })

    # ----- 위탁재고,전시차재고 시트 -----
    consign_sheets = [s for s in wb.sheetnames if '위탁' in s or '전시차' in s]
    for sn in consign_sheets:
        ws = wb[sn]
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = list(row)
                continue
            if not any(row): continue

            # 위탁재고 시트는 컬럼 다름
            i_vin = find_col(header, ['차대 번호', '차대번호', 'VIN'])
            i_com = find_col(header, ['커미션', 'COM'])
            i_mcode = find_col(header, ['모델 코드', '모델코드', 'BM'])
            i_mname = find_col(header, ['모델명', '모델', '차종'])
            i_ext = find_col(header, ['외장', '외색'])
            i_int = find_col(header, ['내장', '내색'])
            i_pdd = find_col(header, ['출고가능일', 'PDD', '출고 가능일'])
            i_branch = find_col(header, ['전시장', '딜러'])
            i_class = find_col(header, ['재고구분', '재고 구분'])
            break

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            if not any(row): continue

            def g(idx):
                return row[idx] if 0 <= idx < len(row) else None

            vin = g(i_vin)
            mname = clean_model(g(i_mname))
            if not mname or not vin: continue

            inv_class_val = clean_model(g(i_class)) or sn  # '위탁재고' or '전시차재고'

            rows.append({
                'date': date_str,
                'vin': str(vin).strip(),
                'com': str(g(i_com)).strip() if g(i_com) else None,
                'model': mname,
                'model_code': str(g(i_mcode)).strip() if g(i_mcode) else None,
                'ext_color': clean_model(g(i_ext)),
                'int_color': clean_model(g(i_int)),
                'pdd': str(g(i_pdd)) if g(i_pdd) else None,
                'prod_date': None,
                'branch': clean_model(g(i_branch)) or '모터원',
                'team': '',
                'salesman': '',
                'customer': '',
                'inv_class': inv_class_val,
                'car_status': '판매 가능',
                'sale_status': '미배정',
                'stock_status': '미배정',
                'vehicle_purpose': '',
                'process_type': '',
                'unsold': '',
                'is_virtual': False,  # 위탁/전시는 실재 차량
                'source': sn,
            })

    return {
        'date': date_str,
        'filename': fp.name,
        'rows': rows,
    }


# ============== 스냅샷 생성 (집계) ==============
def build_snapshot(parsed):
    """파싱 결과 → 모델별 집계 스냅샷.

    G클래스 메인 데이터에서 제외. 단, 별도 g_class_total로는 보존.

    리턴:
      {
        date, total_records, sellable_total, virtual_total, actual_total,
        models: {모델명: {cat, total, virtual, actual, sellable, assigned, motorone}}
        vins: {vin: {model, status, branch, ...}}  # 다음날 비교용
        g_class: { ... 별도 보관 ... }
      }
    """
    rows = parsed['rows']

    # 위탁/전시차는 "일반 재고와 완전 합산" — 별도 처리 없이 그대로 집계
    # 단, VIN 중복 제거 (같은 VIN이 allocation과 위탁에 모두 있을 수 있음)
    seen_vins = set()
    deduped = []
    for r in rows:
        v = r['vin']
        if not v: continue
        if v in seen_vins: continue
        seen_vins.add(v)
        deduped.append(r)

    models = {}
    g_models = {}
    vins_meta = {}

    for r in rows:
        v = r['vin']
        if not v: continue

        is_g = is_g_class(r['model'])
        target = g_models if is_g else models

        if r['model'] not in target:
            target[r['model']] = {
                'cat': categorize(r['model']),
                'total': 0,
                'virtual': 0,
                'actual': 0,
                'sellable': 0,
                'assigned': 0,
                'motorone': 0,  # 모터원 배정재고
                'consign': 0,   # 위탁/전시차
            }
        m = target[r['model']]
        m['total'] += 1
        if r['is_virtual']:
            m['virtual'] += 1
        else:
            m['actual'] += 1
        if r['car_status'] == '판매 가능':
            m['sellable'] += 1
        if r['sale_status'] in ('가계약', '계약 확정', '배정'):
            m['assigned'] += 1
        if '배정재고' in (r['inv_class'] or ''):
            m['motorone'] += 1
        if r['inv_class'] in ('위탁재고', '전시차재고') or r['source'] != 'allocation':
            m['consign'] += 1

        vins_meta[v] = {
            'model': r['model'],
            'cat': categorize(r['model']),
            'is_virtual': r['is_virtual'],
            'is_g': is_g,
            'inv_class': r['inv_class'],
            'sale_status': r['sale_status'],
            'salesman': r['salesman'],
            'customer': r['customer'],
            'branch': r['branch'],
            'pdd': r['pdd'],
        }

    sellable_total = sum(m['sellable'] for m in models.values())
    virtual_total = sum(m['virtual'] for m in models.values())
    actual_total = sum(m['actual'] for m in models.values())

    return {
        'date': parsed['date'],
        'filename': parsed['filename'],
        'total_records': len(deduped),
        'sellable_total': sellable_total,
        'virtual_total': virtual_total,
        'actual_total': actual_total,
        'models': models,
        'g_class': {
            'models': g_models,
            'total': sum(m['total'] for m in g_models.values()),
            'virtual': sum(m['virtual'] for m in g_models.values()),
            'actual': sum(m['actual'] for m in g_models.values()),
            'sellable': sum(m['sellable'] for m in g_models.values()),
        },
        'vins_meta': vins_meta,
    }


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('사용: python3 parse_inventory_v3.py <엑셀파일> [출력json]')
        sys.exit(1)
    fp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None

    parsed = parse_excel(fp)
    snap = build_snapshot(parsed)

    if out:
        # vins_meta는 너무 크므로 별도 파일로 분리 가능
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(snap, f, ensure_ascii=False, indent=2, default=str)
        print(f'저장 완료: {out}')
    else:
        print(json.dumps({k: