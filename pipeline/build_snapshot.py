"""
BenzDream inventory pipeline.

Input: daily xlsx (차량 재고현황)
Output: one JSON snapshot per day under pipeline/snapshots/YYYY-MM-DD.json
Aggregate: pipeline/history.json  (list of snapshots, sorted by date)

Snapshot schema (per day):
{
  "date": "YYYY-MM-DD",
  "sellable_total": int,       # 판매가능 재고 총수 (고객용 노출 대상)
  "assigned_total": int,       # 모터원 배정재고 (가계약)
  "sellable_vins": [...],      # VIN list of sellable pool
  "assigned_vins": [...],      # VIN list of assigned pool
  "assigned_details": {vin: {model, salesman, branch, color, pdd}},
  "models": {
    "<canonical model>": {
      "cat": str,              # 카테고리 (E클래스, GLE, 전동화 EQ, ...)
      "sellable": int,
      "assigned": int,
      "colors": {(ext,int): count},   # 판매가능 기준 색상 조합
    }
  }
}
"""
import json, re, sys, os
from collections import defaultdict, Counter
from pathlib import Path
import openpyxl

# -----------------------
# Model name normalization
# -----------------------
# Reduce raw names to canonical form. Canonical = 4/14 allocation naming.

# Explicit abbreviations -> expansion (applied as whole-word replacements)
ABBR_MAP = [
    (r"\bAV\b", "AVANTGARDE"),
    (r"\bEX\b", "EXCLUSIVE"),
    (r"\b4M\+", "4MATIC+"),
    (r"\b4M\b", "4MATIC"),
    (r"\bHYBRID\b", "Hybrid"),
    (r"\bCOUPE\b", "Coupé"),
    (r"\bCoupe\b", "Coupé"),
]

# Drop-suffix models: 4/10 used longer names, 4/14 dropped "AMG Line" etc.
SUFFIX_DROP = {
    "CLE 200 Cabriolet AMG Line": "CLE 200 Cabriolet",
    "CLE 200 Coupé AMG Line": "CLE 200 Coupé",
    "CLE 450 4MATIC Cabriolet AMG Line": "CLE 450 4MATIC Cabriolet",
    "CLE 450 4MATIC Coupé AMG Line": "CLE 450 4MATIC Coupé",
    "GLB 250 4MATIC AMG Line": "GLB 250 4MATIC",
    "AMG CLA 45 S 4MATIC+ Coupé": "AMG CLA 45 S 4MATIC+ Final Edition Coupé",
    "AMG G 63 MANUFAKTUR": "AMG  G 63 MANUFAKTUR",  # 4/14 canonical uses double space
    "E 350 e 4MATIC EXCLUSIVE with EQ Hybrid": "E 350 e 4MATIC EXCLUSIVE with EQ Hybrid Technology",
    "S 450 4MATIC": "S 450 4MATIC Sedan",
    "S 500 4MATIC": "S 500 4MATIC Sedan long",
    "Maybach S 580": "Maybach S 580 4MATIC",
    "Maybach S 680": "Maybach S 680 4MATIC",
    "E 300 4MATIC AMG Line": "E 300 4MATIC AMG Line",  # identity
    # Short-form wi sheet / prior-day variants
    "GLA 250 4MATIC": "GLA 250 4MATIC AMG Line",
    "E 220 d 4MATIC": "E 220 d 4MATIC EXCLUSIVE",
    "E 350 e 4MATIC EXCLUSIVE": "E 350 e 4MATIC EXCLUSIVE with EQ Hybrid Technology",
    "AMG CLA 45 S 4MATIC+ Final Edition": "AMG CLA 45 S 4MATIC+ Final Edition Coupé",
}

def normalize_model(raw):
    if not raw:
        return None
    s = str(raw).strip()
    # Expand abbreviations first
    for pat, repl in ABBR_MAP:
        s = re.sub(pat, repl, s)
    # Collapse multi-spaces EXCEPT we preserve the canonical "AMG  G 63 MANUFAKTUR" double space
    # Approach: collapse then re-apply drop-map
    s_collapsed = re.sub(r"\s+", " ", s).strip()
    # Explicit suffix drop mapping first
    if s_collapsed in SUFFIX_DROP:
        return SUFFIX_DROP[s_collapsed]
    if s in SUFFIX_DROP:
        return SUFFIX_DROP[s]
    return s_collapsed

# -----------------------
# MB DataCard color code → Korean name
# 위탁재고/전시차재고 시트는 공장 DataCard 코드를 그대로 내려주므로 매핑 필요
# -----------------------
EXT_CODE_MAP = {
    "040": "나이트 블랙",
    "149": "폴라 화이트",
    "183": "마그네타이트 블랙",
    "191": "나이트 블랙",
    "197": "옵시디안 블랙",
    "297": "MANUFAKTUR 가넷 레드 메탈릭",
    "662": "MANUFAKTUR 그라파이트 그레이 마그노",
    "696": "마운틴 그레이",
    "787": "마운틴 그레이 메탈릭",
    "799": "다이아몬드 화이트",
    "817": "하이테크 실버",
    "885": "MANUFAKTUR 하이퍼 블루 마그노",
    "890": "카바나이트 블루",
    "922": "MANUFAKTUR 오팔라이트 화이트 마그노",
    "956": "MANUFAKTUR 알파인 그레이 솔리드",
    "970": "스펙트럴 블루",
    "992": "셀레나이트 그레이",
}
INT_CODE_MAP = {
    "101": "블랙",
    "104": "마키아토 베이지/블랙",
    "105": "블랙/블랙",
    "114": "블랙",
    "118": "시에나 브라운/블랙",
    "121": "블랙",
    "124": "트러플 브라운/블랙",
    "135": "마키아토 베이지/블랙",
    "194": "네바 그레이/블랙",
    "201": "블랙",
    "204": "파워 레드/블랙",
    "205": "마키아토 베이지/블랙",
    "207": "시에나 브라운/블랙",
    "214": "네바 그레이/블랙",
    "215": "네바 그레이/발바오 브라운",
    "221": "마키아토 베이지/마그마 그레이",
    "251": "블랙",
    "511": "마이바흐 익스클루시브 나파 가죽, 블랙",
    "514": "마이바흐 익스클루시브 나파 가죽, 시에나 브라운/블랙",
    "515": "마이바흐 익스클루시브 나파 가죽, 마키아토 베이지/마그마 그레이",
    "651": "아티코 인조 가죽/다이나미카 블랙",
    "654": "아티코 인조 가죽/다이나미카 블랙",
    "671": "AMG 블랙/레드",
    "804": "나파 가죽, 블랙",
    "805": "나파 가죽, 시에나 브라운/블랙",
    "851": "AMG 나파 가죽, 블랙",
    "855": "AMG 나파 가죽, 시에나 브라운/블랙",
    "887": "AMG 나파 가죽, 레드 페퍼/블랙",
    "951": "크리스탈 화이트",
}
def resolve_color(raw, code_map):
    if raw is None: return None
    s = str(raw).strip()
    if not s: return None
    if s.isdigit():
        return code_map.get(s, f"코드 {s}")
    return s

# -----------------------
# Category mapping from 4/14 클래스구분 sheet if possible, else fallback
# -----------------------
CAT_OVERRIDES = {
    # model -> high-level category shown to user
}
def category_from_model(m):
    if not m: return "기타"
    if m.startswith("AMG G 63") or m.startswith("G ") or "G 450" in m or "G 580" in m: return "G클래스"
    if m.startswith("Maybach"): return "Maybach"
    if m.startswith("A "): return "A클래스"
    if m.startswith("AMG A"): return "A클래스"
    if m.startswith("CLA") or m.startswith("AMG CLA"): return "CLA"
    if m.startswith("C ") or m.startswith("AMG C"): return "C클래스"
    if m.startswith("CLE") or m.startswith("AMG CLE"): return "CLE"
    if m.startswith("E ") or m.startswith("AMG E"): return "E클래스"
    if m.startswith("S ") or m.startswith("AMG S"): return "S클래스"
    if m.startswith("EQ"): return "전동화 EQ"
    if m.startswith("GLA"): return "GLA·GLB"
    if m.startswith("GLB") or m.startswith("AMG GLB"): return "GLA·GLB"
    if m.startswith("GLC") or m.startswith("AMG GLC"): return "GLC"
    if m.startswith("GLE") or m.startswith("AMG GLE"): return "GLE"
    if m.startswith("GLS") or m.startswith("AMG GLS"): return "GLS"
    if m.startswith("AMG GT") or m.startswith("AMG SL"): return "AMG GT·SL"
    return "기타"

# -----------------------
# Main snapshot builder
# -----------------------
def build_snapshot(xlsx_path, date_str):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['allocation']
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h:i for i,h in enumerate(hdr) if h}

    # Resolve column positions (handle schema variations)
    c_jg  = idx.get('재고구분')
    c_jg2 = idx.get('재고구분2')
    c_branch = idx.get('배정 전시장')
    c_team = idx.get('팀')
    c_sales = None
    for k in idx:
        if k and 'Sales' in str(k):
            c_sales = idx[k]; break
    c_model = idx.get('모델명')
    c_class1 = idx.get('Class1')
    c_vin = idx.get('차대 번호')
    c_ext = idx.get('외장 색상')
    c_int = idx.get('내장 색상')
    c_pdd = idx.get('차량 출고 가능일(PDD)')

    sellable_vins = set()
    assigned_vins = set()
    assigned_details = {}
    models = {}
    vin_model = {}  # vin -> canonical model name

    def ensure_model(canon):
        if canon not in models:
            models[canon] = {
                "cat": category_from_model(canon),
                "sellable": 0,
                "assigned": 0,
                "colors": defaultdict(int),
            }
        return models[canon]

    for r in rows[1:]:
        jg = r[c_jg] if c_jg is not None else None
        jg2 = r[c_jg2] if c_jg2 is not None else None

        # Skip Virtual VIN (Dummy) in any casing
        if jg2 and 'Virtual' in str(jg2):
            continue
        if jg == '출고':
            continue

        raw_model = r[c_model] if c_model is not None else None
        canon = normalize_model(raw_model)
        if not canon:
            continue
        vin = r[c_vin] if c_vin is not None else None
        ext = resolve_color(r[c_ext] if c_ext is not None else None, EXT_CODE_MAP)
        intc = resolve_color(r[c_int] if c_int is not None else None, INT_CODE_MAP)

        m = ensure_model(canon)

        if jg == '배정재고':
            m["assigned"] += 1
            if vin:
                assigned_vins.add(vin)
                vin_model[vin] = canon
                assigned_details[vin] = {
                    "model": canon,
                    "branch": r[c_branch] if c_branch is not None else None,
                    "team": r[c_team] if c_team is not None else None,
                    "salesman": r[c_sales] if c_sales is not None else None,
                    "ext": ext, "int": intc,
                    "pdd": str(r[c_pdd])[:10] if c_pdd is not None and r[c_pdd] else None,
                }
        else:
            # 전국재고, 전시차재고, 위탁재고 = sellable
            m["sellable"] += 1
            if vin:
                sellable_vins.add(vin)
                vin_model[vin] = canon
            key = f"{ext or '-'}|{intc or '-'}"
            m["colors"][key] += 1

    # Merge separate 위탁/전시차 sheet (avoid VIN duplication with allocation)
    if '위탁재고,전시차재고' in wb.sheetnames:
        ws2 = wb['위탁재고,전시차재고']
        rows2 = list(ws2.iter_rows(values_only=True))
        hdr2 = rows2[0]
        i2 = {h:i for i,h in enumerate(hdr2) if h}
        cv = i2.get('VIN')
        cm = i2.get('Model')
        cc = i2.get('Color')
        cu = i2.get('Upholstery')
        cjg = i2.get('재고구분')
        for r in rows2[1:]:
            jg = r[cjg] if cjg is not None else None
            if jg not in ('위탁재고', '전시차재고'):
                continue
            vin = r[cv] if cv is not None else None
            if vin and vin in sellable_vins:
                continue  # already counted in allocation
            if vin and vin in assigned_vins:
                continue
            raw = r[cm] if cm is not None else None
            canon = normalize_model(raw)
            if not canon: continue
            m = ensure_model(canon)
            m["sellable"] += 1
            if vin:
                sellable_vins.add(vin)
                vin_model[vin] = canon
            ext2 = resolve_color(r[cc] if cc is not None else None, EXT_CODE_MAP) or '-'
            int2 = resolve_color(r[cu] if cu is not None else None, INT_CODE_MAP) or '-'
            key = f"{ext2}|{int2}"
            m["colors"][key] += 1

    # Finalize
    for canon, data in models.items():
        data["colors"] = dict(data["colors"])

    snapshot = {
        "date": date_str,
        "sellable_total": len(sellable_vins),
        "assigned_total": len(assigned_vins),
        "sellable_vins": sorted(sellable_vins),
        "assigned_vins": sorted(assigned_vins),
        "assigned_details": assigned_details,
        "models": models,
        "vin_model": vin_model,
    }
    return snapshot

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: build_snapshot.py <xlsx> <YYYY-MM-DD>")
        sys.exit(1)
    snap = build_snapshot(sys.argv[1], sys.argv[2])
    out_dir = Path(__file__).parent / "snapshots"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"{sys.argv[2]}.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    print(f"✓ saved {out_path}")
    print(f"  sellable={snap['sellable_total']}, assigned={snap['assigned_total']}, models={len(snap['models'])}")
